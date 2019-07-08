# sendDuesReminders.py - Sends emails based on their status in spreadsheet.

import openpyxl, smtplib, sys
from email.mime.text import MIMEText

# Open the spreadsheet and get the latest dues status.
wb = openpyxl.load_workbook('duesRecords.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

lastCol = sheet.max_column
latestMonth = sheet.cell(row=1, column=lastCol).value

unpaidMembers = {}

# Check each member's payment status
for r in range(2, sheet.max_row + 1):
    payment = sheet.cell(row=r, column=lastCol).value
    if payment != 'paid':
        name = sheet.cell(row=r, column=1).value
        email = sheet.cell(row=r, column=2).value
        if email != None:
            unpaidMembers[name] = email

# Log in to email account.
smtpObj = smtplib.SMTP('smtp.gmail.com', 587)
smtpObj.starttls()
# smtpObj.login('pythonlink@gmail.com', sys.argv[1])
smtpObj.login('pythonlink@gmail.com', 'ikwanmgexipeygva')

# Send out reminder emails.
for name, email in unpaidMembers.items():
    body = '\n我说内个{}啊,\n\n' \
           ' {}的党费你想拖到什么时候啊，赶紧交！明天不把帐结清我卸了你家蟑螂的腿！！！\n\n ' \
           '大咖党党支部 '.format(name, latestMonth)
    print('Sending email to {}...'.format(email))
#
    msg = MIMEText(body)
    msg['From'] = 'pythonlink@gmail.com'
    msg['To'] = email
    msg['Subject'] = '拖欠的人是可耻的'
    sendmailStatus = smtpObj.send_message(msg)
#
    if sendmailStatus != {}:
        print('There was a problem sending email to {}: {}'.format(email, sendmailStatus))
smtpObj.quit()

import sendMessage

message = '赶紧把大咖党党费交了！！！'
sendMessage.textMyself(message)
