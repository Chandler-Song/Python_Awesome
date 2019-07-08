import openpyxl
from openpyxl.utils import get_column_letter

# wb = openpyxl.Workbook()
# sheet = wb.active
#
# # change the name of the sheet
# print(sheet.title)
# sheet.title = 'Happy2017'
# print(wb.get_sheet_names())
#
# wb.create_sheet(index=0, title='First Sheet')
# wb.create_sheet(index=1, title='Middle Sheet')
# print(wb.get_sheet_names())
#
# wb.remove_sheet(wb.get_sheet_by_name('Middle Sheet'))
# print(wb.get_sheet_names())
#
# wb.save('temp1.xlsx')

# write values to cells
wb = openpyxl.Workbook()
sheet = wb.active

sheet['A1'] = 'Hello python'
print(sheet['A1'].value)

ws1 = wb.create_sheet('range names')
for row in range(1, 40):
	ws1.append(range(17))

ws2 = wb.create_sheet('List')
rows = [
	['Number', 'Batch 1', 'Batch2'],
	[2, 40, 30],
	[3, 40, 25],
	[4, 50, 30],
	[5, 30, 10],
	[6, 40, 30],
	[7, 78, 52],
]
for row in rows:
	ws2.append(row)

ws3 = wb.create_sheet(title='Data')
for row in range(5, 30):
	for col in range(15, 54):
		ws3.cell(column=col, row=row, value=get_column_letter(col))
print(ws3['AA10'].value)

wb.save(filename='empty_book.xlsx')
