# Corret cost in produceSale spreadsheet
import openpyxl

# The product types and their updated prices
PRICE_UPDATE = {
	'Garlic': 3.17,
    'Celery':1.19,
    'Lemon': 1.27
}

wb = openpyxl.load_workbook('produceSales.xlsx')
ws = wb.get_sheet_by_name('Sheet')

# loop through the rows and update the prices, skip the first row
for rowNum in range(2, ws.max_row+1):

	productName = ws.cell(row=rowNum, column=1).value

	if productName in PRICE_UPDATE:
		ws.cell(row=rowNum, column=2).value = PRICE_UPDATE[productName]

wb.save('updateProduceSales.xlsx')